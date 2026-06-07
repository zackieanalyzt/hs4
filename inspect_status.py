import openpyxl

wb = openpyxl.load_workbook('การประเมิน HS4.xlsx', data_only=True)
target_sheets = ['2. ด้านการบริการสุขภาพ', 'ด้านที่ 5 ความปลอดภัย', 'ด้านที่ 6 เครื่องมืออุปกรณ์ทางก', 'ด้านที่ 7 ระบบสนับสนุนบริการที่', 'ด้านที่ 8 สุขศึกษาและพฤติกรรมสุ']

print(f"Investigating status columns in: {target_sheets}")

for sheet_name in target_sheets:
    if sheet_name not in wb.sheetnames:
        print(f"Sheet {sheet_name} not found!")
        continue
    
    sheet = wb[sheet_name]
    print(f"\n--- {sheet_name} ---")
    
    # Read first 10 rows to find headers and data
    for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True)):
        print(f"Row {i+1}: {row[:10]}")

print("\n--- OverAll Sheet ---")
overall = wb['OverAll']
for i, row in enumerate(overall.iter_rows(min_row=1, max_row=20, values_only=True)):
    print(f"Row {i+1}: {row[:10]}")
