import openpyxl

wb = openpyxl.load_workbook('การประเมิน HS4.xlsx', data_only=True)
target_sheets = ['2. ด้านการบริการสุขภาพ', 'ด้านที่ 5 ความปลอดภัย', 'ด้านที่ 6 เครื่องมืออุปกรณ์ทางก', 'ด้านที่ 7 ระบบสนับสนุนบริการที่', 'ด้านที่ 8 สุขศึกษาและพฤติกรรมสุ']

for sheet_name in target_sheets:
    sheet = wb[sheet_name]
    print(f"\n--- Sheet: {sheet_name} ---")
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
        if row[0] == 'ข้อที่':
            print(f"Header at Row {row_idx+1}: {row}")
            # Show first 5 data rows
            for d_idx in range(row_idx+1, min(row_idx+6, sheet.max_row)):
                data_row = [cell.value for cell in sheet[d_idx+1]]
                print(f"  Data Row {d_idx+1}: {data_row[:10]}")
            break
