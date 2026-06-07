import openpyxl
import json
import os

def extract_data(file_path):
    if not os.path.exists(file_path):
        print(f"Error: {file_path} not found.")
        return None

    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    # Configuration for specific sheets
    target_sheets = [
        '2. ด้านการบริการสุขภาพ', 
        'ด้านที่ 5 ความปลอดภัย', 
        'ด้านที่ 6 เครื่องมืออุปกรณ์ทางก', 
        'ด้านที่ 7 ระบบสนับสนุนบริการที่', 
        'ด้านที่ 8 สุขศึกษาและพฤติกรรมสุ'
    ]
    
    dashboard_data = []

    for sheet_name in target_sheets:
        if sheet_name not in wb.sheetnames:
            continue
            
        sheet = wb[sheet_name]
        status_counts = {'ยังไม่ได้แก้ไข': 0, 'แก้ไขแล้ว': 0, 'ไม่แก้ไข ยืนยันเดิม': 0}
        
        header_row_idx = -1
        col_idx = 5 
        
        for idx, row in enumerate(sheet.iter_rows(max_row=30, values_only=True)):
            if row[0] == 'ข้อที่':
                header_row_idx = idx + 1
                for i, cell in enumerate(row):
                    if cell and 'การแก้ไข' in str(cell):
                        col_idx = i
                break
        
        if header_row_idx != -1:
            for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
                if not row[0] or (isinstance(row[0], str) and not row[0][0].isdigit()):
                    continue
                
                status = row[col_idx] if len(row) > col_idx else None
                if status in status_counts:
                    status_counts[status] += 1

        dashboard_data.append({
            "aspect": sheet_name,
            "status_counts": status_counts
        })

    return {"aspects": dashboard_data}

def generate_html(data):
    html_template = """
<!DOCTYPE html>
<html lang="th">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>HS4 Correction Progress</title>
    <script src="https://cdn.jsdelivr.net/npm/echarts@5.4.3/dist/echarts.min.js"></script>
    <link href="https://fonts.googleapis.com/css2?family=Sarabun:wght@300;400;700&display=swap" rel="stylesheet">
    <style>
        * { box-sizing: border-box; margin: 0; padding: 0; }
        body { 
            font-family: 'Sarabun', sans-serif; 
            background-color: #f1f5f9; 
            color: #1e293b;
            padding: 40px 20px;
            line-height: 1.5;
        }
        .container { 
            max-width: 1200px; 
            margin: 0 auto; 
        }
        h1 { 
            text-align: center; 
            font-size: 28px;
            margin-bottom: 40px; 
            color: #0f172a;
        }
        .grid { 
            display: grid; 
            grid-template-columns: repeat(auto-fill, minmax(350px, 1fr)); 
            gap: 24px; 
        }
        .card { 
            background: #ffffff; 
            border-radius: 16px; 
            padding: 24px; 
            box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1), 0 2px 4px -2px rgba(0, 0, 0, 0.1);
            display: flex; 
            flex-direction: column;
            height: 450px; /* Fixed height to prevent layout shifting */
        }
        .card-title { 
            font-weight: 700; 
            font-size: 18px; 
            color: #334155; 
            text-align: center; 
            margin-bottom: 20px;
            height: 54px;
            display: flex;
            align-items: center;
            justify-content: center;
            border-bottom: 1px solid #e2e8f0;
            padding-bottom: 10px;
        }
        .pie-container { 
            flex-grow: 1;
            width: 100%;
        }
        .no-data { 
            display: flex;
            align-items: center;
            justify-content: center;
            height: 100%;
            color: #94a3b8; 
            font-style: italic; 
        }
    </style>
</head>
<body>
    <div class="container">
        <h1>สัดส่วนการแก้ไขมาตรฐาน HS4 (ด้านที่ 2, 5, 6, 7, 8)</h1>
        <div class="grid" id="charts-grid"></div>
    </div>

    <script>
        const data = DATA_PLACEHOLDER;
        const grid = document.getElementById('charts-grid');

        data.aspects.forEach((item, idx) => {
            const card = document.createElement('div');
            card.className = 'card';
            
            const total = Object.values(item.status_counts).reduce((a, b) => a + b, 0);
            
            let contentHtml = `<div class="card-title">${item.aspect}</div>`;
            if (total > 0) {
                contentHtml += `<div id="pie-${idx}" class="pie-container"></div>`;
            } else {
                contentHtml += `<div class="no-data">ไม่มีข้อมูลที่ต้องแก้ไข</div>`;
            }
            
            card.innerHTML = contentHtml;
            grid.appendChild(card);

            if (total > 0) {
                const pChart = echarts.init(document.getElementById(`pie-${idx}`));
                pChart.setOption({
                    tooltip: { trigger: 'item', formatter: '{b}: {c} ({d}%)' },
                    legend: { 
                        bottom: '0', 
                        left: 'center', 
                        textStyle: { fontSize: 12 },
                        itemWidth: 10,
                        itemHeight: 10
                    },
                    series: [{
                        type: 'pie',
                        radius: ['45%', '70%'],
                        center: ['50%', '45%'],
                        avoidLabelOverlap: true,
                        itemStyle: { borderRadius: 10, borderColor: '#fff', borderWidth: 2 },
                        label: { show: true, position: 'outside', formatter: '{d}%', fontSize: 12 },
                        emphasis: { label: { show: true, fontSize: 16, fontWeight: 'bold' } },
                        data: [
                            { value: item.status_counts['ยังไม่ได้แก้ไข'], name: 'ยังไม่ได้แก้ไข', itemStyle: {color: '#ef4444'} },
                            { value: item.status_counts['แก้ไขแล้ว'], name: 'แก้ไขแล้ว', itemStyle: {color: '#22c55e'} },
                            { value: item.status_counts['ไม่แก้ไข ยืนยันเดิม'], name: 'ไม่แก้ไข ยืนยันเดิม', itemStyle: {color: '#3b82f6'} }
                        ]
                    }]
                });
            }
        });

        window.addEventListener('resize', () => {
            data.aspects.forEach((_, idx) => {
                const chart = echarts.getInstanceByDom(document.getElementById(`pie-${idx}`));
                if (chart) chart.resize();
            });
        });
    </script>
</body>
</html>
"""
    json_data = json.dumps(data, ensure_ascii=False)
    final_html = html_template.replace('DATA_PLACEHOLDER', json_data)
    
    with open('dashboard.html', 'w', encoding='utf-8') as f:
        f.write(final_html)
    print("Dashboard updated and layout fixed: dashboard.html")

if __name__ == "__main__":
    extracted_data = extract_data('การประเมิน HS4.xlsx')
    if extracted_data:
        generate_html(extracted_data)
