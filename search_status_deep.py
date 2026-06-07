import openpyxl

wb = openpyxl.load_workbook('การประเมิน HS4.xlsx', data_only=True)
target_sheets = ['2. ด้านการบริการสุขภาพ', 'ด้านที่ 5 ความปลอดภัย', 'ด้านที่ 6 เครื่องมืออุปกรณ์ทางก', 'ด้านที่ 7 ระบบสนับสนุนบริการที่', 'ด้านที่ 8 สุขศึกษาและพฤติกรรมสุ', 'OverAll']

for sheet_name in target_sheets:
    sheet = wb[sheet_name]
    print(f"\n--- Searching in: {sheet_name} ---")
    
    # Find headers
    found_header = False
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
        if any(keyword in str(cell) for cell in row if cell for keyword in ['การแก้ไข', 'สถานะ']):
            print(f"Header found at Row {row_idx+1}: {row}")
            found_header = True
            
            # If we found a header, let's look at the data below it
            # and count unique values in the "correction" column
            # Assume it's the last column or look for 'การแก้ไข' specifically
            col_idx = -1
            for i, cell in enumerate(row):
                if cell and 'การแก้ไข' in str(cell):
                    col_idx = i
                    break
            
            if col_idx != -1:
                counts = {}
                for data_row in sheet.iter_rows(min_row=row_idx+2, values_only=True):
                    val = data_row[col_idx]
                    if val:
                        counts[val] = counts.get(val, 0) + 1
                print(f"Status Counts: {counts}")
            break
            
    if not found_header:
        print("No 'การแก้ไข' header found.")
