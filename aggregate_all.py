import openpyxl

wb = openpyxl.load_workbook('การประเมิน HS4.xlsx', data_only=True)
all_aspect_sheets = [s for s in wb.sheetnames if s != 'OverAll']

for sheet_name in all_aspect_sheets:
    sheet = wb[sheet_name]
    found_col = False
    for row in sheet.iter_rows(min_row=1, max_row=20, values_only=True):
        if 'ข้อที่' in str(row[0]):
            col_idx = -1
            for i, cell in enumerate(row):
                if cell and 'การแก้ไข' in str(cell):
                    col_idx = i
                    break
            if col_idx == -1:
                # If no header, check if column 5 or 6 has status values
                col_idx = 5 # Default guess based on observation
            
            print(f"Sheet: {sheet_name}, Using Column Index: {col_idx}")
            
            counts = {}
            for d_row in sheet.iter_rows(min_row=1, values_only=True):
                # Skip rows that are headers or summary
                if len(d_row) > col_idx and d_row[col_idx] in ['ยังไม่ได้แก้ไข', 'แก้ไขแล้ว', 'ไม่แก้ไข ยืนยันเดิม']:
                    val = d_row[col_idx]
                    counts[val] = counts.get(val, 0) + 1
                elif len(d_row) > col_idx-1 and d_row[col_idx-1] == 'N/A' and d_row[col_idx] is None:
                    counts['ไม่ต้องแก้ไข'] = counts.get('ไม่ต้องแก้ไข', 0) + 1
            
            print(f"  Counts: {counts}")
            found_col = True
            break
    if not found_col:
        print(f"Sheet: {sheet_name}, No status data found.")
